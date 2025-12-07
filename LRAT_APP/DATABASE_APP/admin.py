from django.contrib import admin
from django import forms
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import Group
from django.http import HttpResponse
from django.db.models import Count, Sum, Q
from decimal import Decimal
from .models import CustomUser, Software, SoftwareOption, Subscription, DatabaseGroup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
# Register your models here.

admin.site.unregister(Group)

class GroupUserInline(admin.TabularInline):
    model = CustomUser.groups.through
    fk_name = "group"
    ordering = ("customuser__username",)
    verbose_name_plural = "Users"
    fields = ("username_display", "first_name_display", "last_name_display", "user_email")
    readonly_fields = ("username_display", "first_name_display", "last_name_display", "user_email")

    def has_add_permission(self, request, obj=None):
        return False

    @admin.display(description="Username")
    def username_display(self, instance):
        return instance.customuser.username

    @admin.display(description="First Name")
    def first_name_display(self, instance):
        return instance.customuser.first_name

    @admin.display(description="Last Name")
    def last_name_display(self, instance):
        return instance.customuser.last_name
    
    @admin.display(description="Email")
    def user_email(self, obj):
        return obj.customuser.email
    
@admin.register(DatabaseGroup)
class GroupAdmin(admin.ModelAdmin):
    list_display = ("name", "user_count")
    ordering = ("name",)
    inlines = (GroupUserInline,)
    search_fields = ("name",)

    @admin.display(description="Users")
    def user_count(self, obj):
        return obj.user_set.count()

def _customuser_group_str(self):
    user = getattr(self, "customuser", None)
    if user:
        return ""
    return ""

CustomUser.groups.through.__str__ = _customuser_group_str

class SubscriptionInline(admin.TabularInline):
    model = Subscription
    extra = 0
    fields = ("software", "option", "currently_used", "renew", "total_cost")
    readonly_fields = ("software", "option", "total_cost")
    can_delete = False
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False

@admin.register(CustomUser)
class CustomUserAdmin(UserAdmin):
    list_display = UserAdmin.list_display + ("is_active", "group_names",)
    search_fields = ("username", "email", "first_name", "last_name")
    list_filter = ("groups", "is_staff", "is_superuser", "is_active")
    ordering = ("last_name",)
    inlines = (SubscriptionInline,)

    @admin.display(description="Groups")
    def group_names(self, obj):
        return ", ".join([group.name for group in obj.groups.all()])

class SoftwareOptionInline(admin.TabularInline):
    model = SoftwareOption
    extra = 1
    show_change_link = True
    fields = ("name", "cost", "is_active")

@admin.register(Software)
class SoftwareAdmin(admin.ModelAdmin):
    list_display = ("subscription_name", "formatted_cost", "term_display", "is_active", "license_start", "license_end")
    search_fields = ("subscription_name", "description")
    list_filter = ("is_active",)
    ordering = ("subscription_name",)
    inlines = [SoftwareOptionInline]

    @admin.display(description="Cost", ordering="cost")
    def formatted_cost(self, obj):
        return f"${obj.cost:,.2f}"
    
    @admin.display(description="Term", ordering="term")
    def term_display(self, obj):
        if obj.term == 1:
            return "1 month"
        return f"{obj.term} months"

def export_subscription_report(modeladmin, request, queryset):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subscription Report"
    
    #header style
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    #column headers
    headers = [
        "Software Name",
        "Currently Subscribed Users",
        "Users Planning to Renew",
        "Current Quarter Total Cost",
        "Next Quarter Estimated Cost"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    #get all software
    software_list = Software.objects.all().order_by('subscription_name')
    
    row_num = 2
    total_current_cost = Decimal('0.00')
    total_next_quarter_cost = Decimal('0.00')

    #get count of currently subscribed users
    for software in software_list:
        current_subs = Subscription.objects.filter(
            software=software,
            currently_used=True
        )
        current_users_count = current_subs.count()
        
        #get count of users planning to renew
        renew_count = current_subs.filter(renew=True).count()
        
        #calculate current quarter total cost
        current_cost = current_subs.aggregate(
            total=Sum('total_cost')
        )['total'] or Decimal('0.00')
        
        # calculate next quarter estimated cost via renewals
        next_quarter_cost = current_subs.filter(renew=True).aggregate(
            total=Sum('total_cost')
        )['total'] or Decimal('0.00')
        
        ws.cell(row=row_num, column=1, value=software.subscription_name)
        ws.cell(row=row_num, column=2, value=current_users_count)
        ws.cell(row=row_num, column=3, value=renew_count)
        ws.cell(row=row_num, column=4, value=float(current_cost))
        ws.cell(row=row_num, column=5, value=float(next_quarter_cost))
        
        ws.cell(row=row_num, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row_num, column=5).number_format = '"$"#,##0.00'
        
        total_current_cost += current_cost
        total_next_quarter_cost += next_quarter_cost
        
        row_num += 1
    
    # totals
    totals_row = row_num + 1
    ws.cell(row=totals_row, column=1, value="TOTAL")
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    ws.cell(row=totals_row, column=4, value=float(total_current_cost))
    ws.cell(row=totals_row, column=4).number_format = '"$"#,##0.00'
    ws.cell(row=totals_row, column=4).font = Font(bold=True)
    ws.cell(row=totals_row, column=5, value=float(total_next_quarter_cost))
    ws.cell(row=totals_row, column=5).number_format = '"$"#,##0.00'
    ws.cell(row=totals_row, column=5).font = Font(bold=True)
    
    # auto adjust column width (https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    #create HTTP response to trigger file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'subscription_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

export_subscription_report.short_description = "Export subscription report to Excel"

@admin.register(Subscription)
class SubscriptionAdmin(admin.ModelAdmin):
    class SubscriptionForm(forms.ModelForm):
        renew = forms.TypedChoiceField(
            label="Renew",
            choices=(("True", "Renew subscription"), ("False", "Do not renew")),
            coerce=lambda value: value == "True",
            widget=forms.RadioSelect,
            empty_value="False",
        )

        class Meta:
            model = Subscription
            exclude = ("total_cost", "currently_used", "date_expired", )
        
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.fields["renew"].initial = "True" if self.instance.renew else "False"
            option_field = self.fields.get("option")
            if option_field:
                option_field.queryset = SoftwareOption.objects.none()

                if "software" in self.data:
                    try:
                        software_id = int(self.data.get("software"))
                        option_field.queryset = SoftwareOption.objects.filter(
                            software_id=software_id
                        )
                    except (TypeError, ValueError):
                        pass
                elif self.instance.pk and self.instance.software_id:
                    option_field.queryset = self.instance.software.options.all()

    list_display = ("user", "software", "option", "currently_used", "renew", "total_cost_display", "date_subscribed_display", "date_expired_display")
    search_fields = ("user__username", "user__email", "software__subscription_name", "option__name")
    list_filter = ("currently_used", "renew", "date_subscribed", "date_expired")
    ordering = ("user",)
    form = SubscriptionForm
    actions = [export_subscription_report]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related("user", "software", "option")
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.currently_used = True
            obj.date_expired = obj.software.license_end or obj.date_expired
            obj.renew = False
        super().save_model(request, obj, form, change)

    @admin.display(description="Total Cost", ordering="total_cost")
    def total_cost_display(self, obj):
        if obj.total_cost is None:
            obj.total_cost = obj.compute_total_cost()
            obj.save(update_fields=["total_cost"])
        return f"${obj.total_cost:,.2f}"
    
    @admin.display(description="Date Subscribed")
    def date_subscribed_display(self, obj):
        return obj.date_subscribed.strftime("%b %d, %Y")        
    
    @admin.display(description="Date Expired")
    def date_expired_display(self, obj):
        return obj.date_expired.strftime("%b %d, %Y") 